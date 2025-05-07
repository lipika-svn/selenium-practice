import openpyxl
book=openpyxl.load_workbook("C:\\PythonProject1\\python Selenium\\Data\\testdata.xlsx")
sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value='RAHUL'
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
